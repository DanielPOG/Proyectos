from django.shortcuts import render, redirect,get_object_or_404
from fichas.models import FichasXAprendiz
from django.contrib.auth.decorators import login_required
import openpyxl
from django.http import HttpResponse
from usuarios.models import Usuarios
from fichas.models import FichasXAprendiz, Fichas
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.db import transaction
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# Create your views here.
def carnet(request):
    return render(request, 'carnet.html')

@login_required
def carnet_view(request):
    if request.user.is_authenticated:
        user = request.user  # El usuario ya está autenticado, no es necesario hacer otra consulta
        print(user.estado)
        if user.estado == 1:
            # Verificar si el usuario tiene datos faltantes en el campo 'rh'
            if user.rh is None and user.foto is None or user.foto =='' :
                context = {'error': 'DATOS FALTANTES: SOLICITA LLENAR DATOS'}
                return render(request, 'pages_user/login_user.html', context)
            
            else:
                # Obtener las fichas en las que está inscrito el usuario
                try:
                    fichas = FichasXAprendiz.objects.get(aprendiz=user)  # Suponiendo que "aprendiz" es el campo que relaciona FichasXAprendiz con Usuarios
                    ficha_numero = fichas.ficha  # Acceder a la ficha relacionada (asumiendo que "ficha" es el campo de relación)
                    
                    context = {
                        'ficha_numero': ficha_numero
                    }
                    return render(request, 'carnet.html', context)

                except FichasXAprendiz.DoesNotExist:
                    # Si el aprendiz no está inscrito en ninguna ficha, manejar el caso
                    context = {'error': 'No tienes una ficha registrada.'}
                    return render(request, 'pages_user/login_user.html', context)
        else:
            context = {'error': 'Su carnet esta Inactivo'}
            return render(request, 'pages_user/login_user.html', context)

    else:
        return redirect('login_user')  # Si no está autenticado, redirigir a la página de login
    
def carnet_admin(request):
    user= request.user
    if user.rh is None and user.foto is None or user.foto =='' :
            context = {'error': 'DATOS FALTANTES:  LLENAR DATOS'}
            return render(request, 'pages/interfaz_index.html', context)
    context={'user':user}
    return render(request ,'carnet_admin.html')

def carnet_index(request):
    try:
        # Obtener aprendices con ficha
        carnets = FichasXAprendiz.objects.select_related("aprendiz", "ficha").all()

        # Obtener instructores (pueden o no estar en Fichas como líder)
        instructores = Usuarios.objects.filter(rol__tipo_rol="Instructor")

        # Unir listas de aprendices con ficha e instructores
        usuarios = list(carnets) + list(instructores)

        # Contar total de usuarios activos
        total_carnets = Usuarios.objects.filter(estado=1).count()


        # Paginación (10 registros por página)
        paginator = Paginator(usuarios, 10)
        page_number = request.GET.get("page", 1)
        try:
            page_obj = paginator.page(page_number)
        except Exception as e:
            print(f"Error en paginación: {e}")
            return JsonResponse({"error": str(e)}, status=400)

        # Si es una petición AJAX, devolver JSON con los datos
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            data = []
            for usuario in page_obj:
                if isinstance(usuario, FichasXAprendiz):  # Aprendiz con ficha
                    data.append({
                        "ficha": usuario.ficha.numero_ficha,
                        "num_doc": usuario.aprendiz.num_doc,
                        "nombre": usuario.aprendiz.nombre,
                        "apellido": usuario.aprendiz.apellido,
                        "estado": usuario.aprendiz.estado,
                        "rol": str(usuario.aprendiz.rol),
                    })
                else:  # Instructor
                    ficha_asociada = Fichas.objects.filter(lider_ficha=usuario).first()
                    data.append({
                        "ficha": ficha_asociada.numero_ficha if ficha_asociada else "N/A",
                        "num_doc": usuario.num_doc,
                        "nombre": usuario.nombre,
                        "apellido": usuario.apellido,
                        "estado": usuario.estado,
                        "rol": str(usuario.rol),
                    })

            return JsonResponse({
                "data": data,
                "has_next": page_obj.has_next(),
                "has_prev": page_obj.has_previous(),
                "total_carnets": total_carnets
            })

        # Renderizar la vista normal con los datos
        return render(request, "registro/carnet_index.html", {
            "carnets": page_obj,
            "total_carnets": total_carnets
        })

    except Exception as e:
        print(f"Error general: {e}")
        return JsonResponse({"error": "Error inesperado en el servidor"}, status=500)

def estado_carnet(request, id):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    try:
        with transaction.atomic():  # Asegura que el cambio se haga correctamente
            user = get_object_or_404(Usuarios, num_doc=id)
            user.estado = 0 if user.estado == 1 else 1
            user.save()

        # Contar cuántos carnets están activos
        total_carnets = Usuarios.objects.filter(estado=1).count()

        return JsonResponse({
            "success": True,
            "nuevo_estado": user.estado,
            "total_carnets": total_carnets
        })
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def exportar_carnets_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Carnets"

    # Encabezados
    headers = ["Número de Ficha", "Documento", "Nombre", "Estado", "Rol"]
    sheet.append(headers)

    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal="center", vertical="center")

    carnets = FichasXAprendiz.objects.select_related('ficha', 'aprendiz')
    instructores = Usuarios.objects.filter(rol__tipo_rol="Instructor")

    row_num = 2  # Iniciar en la segunda fila

    # Agregar aprendices con ficha
    for carnet in carnets:
        estado_texto = "Activo" if carnet.aprendiz.estado == 1 else "Inactivo"  
        sheet.append([
            carnet.ficha.numero_ficha,
            carnet.aprendiz.num_doc,
            f"{carnet.aprendiz.nombre} {carnet.aprendiz.apellido}",
            estado_texto,
            "Aprendiz"
        ])
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1

    # Agregar instructores con y sin ficha
    for instructor in instructores:
        ficha_asociada = Fichas.objects.filter(lider_ficha=instructor).first()
        ficha_numero = ficha_asociada.numero_ficha if ficha_asociada else "N/A"
        estado_texto = "Activo" if instructor.estado == 1 else "Inactivo"

        sheet.append([
            ficha_numero,
            instructor.num_doc,
            f"{instructor.nombre} {instructor.apellido}",
            estado_texto,
            "Instructor"
        ])
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.border = border_style
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1

    # Ajustar el ancho de las columnas automáticamente
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openpyxl")
    response["Content-Disposition"] = 'attachment; filename="carnets.xlsx"'
    workbook.save(response)
    return response
